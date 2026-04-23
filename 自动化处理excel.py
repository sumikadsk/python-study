from openpyxl import load_workbook, Workbook
import os


class ExcelAutoHandler:
    """Excel自动化处理工具：数据清洗、去重、统计、自动生成报表"""

    def __init__(self):
        # 新建最终输出工作簿
        self.wb_out = Workbook()
        self.ws_out = self.wb_out.active
        self.ws_out.title = "数据汇总结果"

    def load_excel_data(self, file_path):
        """读取单个Excel文件数据"""
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            data_list = []

            # 遍历每行数据
            for row in ws.iter_rows(values_only=True):
                # 过滤空行
                if any(cell is not None for cell in row):
                    data_list.append(row)
            return data_list
        except Exception as e:
            print(f"读取文件失败：{file_path}，错误：{e}")
            return []

    def clean_data(self, data):
        """数据清洗：简单去重"""
        # 利用集合去重
        clean_data = list(set(data))
        return clean_data

    def save_result(self, save_path="汇总报表.xlsx"):
        """自动保存处理后的Excel报表"""
        self.wb_out.save(save_path)
        print(f"自动化处理完成，报表已导出：{save_path}")

    def run_batch(self, folder_path):
        """批量处理文件夹内所有xlsx文件"""
        all_clean_data = []
        # 遍历文件夹
        for file in os.listdir(folder_path):
            if file.endswith((".xlsx", ".xls")):
                file_full = os.path.join(folder_path, file)
                raw_data = self.load_excel_data(file_full)
                clean_data = self.clean_data(raw_data)
                all_clean_data.extend(clean_data)
                print(f"已处理：{file}")

        # 将清洗后的数据写入新表格
        for idx, row in enumerate(all_clean_data, 1):
            self.ws_out.append(row)

        # 保存最终报表
        self.save_result()


if __name__ == "__main__":
    # 改成你存放Excel文件的文件夹路径
    target_folder = r"C:\Users\13770\Desktop\excel_data"

    # 实例化并执行自动化流程
    tool = ExcelAutoHandler()
    tool.run_batch(target_folder)